#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
產生 src/brand_map.json —— 口罩貨號(B開頭) 的「3碼公司碼 → 品牌」對照。

貨號結構（2026-09-05 使用者確認）：B + 3碼公司碼 + 1碼款式碼(1=平面/2=立體) + 尺寸 + 序號。
同一個公司碼底下所有款式碼都算同一個品牌（例：B0041/B0042/B0043 都是郡昱；
B0091/B0092 都是昌明——安心罩護是昌明的立體款分身，不能拆成兩個品牌）。

做法：掃蝦皮獲利計算表的分類分頁（📦 A保健食品/📦 B醫療口罩…；「📦 商品列表」本身
的商品名稱欄常年是公式格讀不到快取值，別用），每個 3 碼公司碼取品名裡的品牌關鍵字、
多數決。再套 OVERRIDES 修正混雜/牛強分類錯的碼。輸出可手改的 JSON。

平常不用跑；只有蝦皮獲利表新增品牌、或發現分類怪怪時，跑一次再手修 JSON。
"""
import openpyxl, re, collections, json, os

HOME = os.path.expanduser("~")
SHOPEE = os.path.join(HOME, "我的雲端硬碟/蝦皮獲利計算表/蝦皮獲利計算表_含稅版_updated.xlsx")
OUT = os.path.abspath(os.path.expanduser(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "brand_map.json")))

# 門市會叫貨的口罩品牌關鍵字（順序＝優先序，長/獨特的放前面）
BRANDS = [
    "安心罩護", "鼻恩恩", "幸福物語", "吉伊卡哇", "健康天使", "愛貝恩", "舒膚康", "挺立舒",
    "郡昱", "興安", "昌明", "匠心", "中衛", "聚泰", "德冠", "凱上", "凱馺", "億宏", "上好",
    "星業", "佑合", "佑和", "水舞", "華淨", "艾爾絲", "明基", "盛籐", "天心", "新寵兒",
    "睿昱", "摩戴舒", "活力安", "艾可兒", "丞威", "沐慷", "康丞", "索菲亞", "萊潔", "月池",
    "優美特", "3M", "宇宙", "卡娜赫拉", "大甲媽", "MissMix", "怡安", "荷康", "BNN", "YOKU",
    "A.O.K",
]

# 混雜碼手動修正（多數決判不準、或票數接近但知道正確答案的）
# key = 3碼公司碼前綴（"B"+3位數字，例如 "B004"），不是完整4碼
OVERRIDES = {
    "B003": "幸福物語",   # 幸福物語/艾爾絲/吉伊卡哇/迪士尼聯名都是同一組(明基幸福物語)賣場
    "B004": "郡昱",
    "B005": "舒膚康",
    "B006": "匠心",
    "B007": "德冠",       # DG / 大甲媽 都是德冠
    "B008": "水舞",
    "B009": "昌明",       # 2026-09-05 使用者確認：安心罩護就是昌明的立體款，同一家公司
    "B010": "凱馺",
    "B011": "上好",
    "B012": "MissMix",
    "B013": "盛籐",       # 盛籐/天心/新寵兒同一組賣場，多數是盛籐
    "B014": "凱上",
    "B017": "億宏",
    "B019": "星業",
    "B020": "億宏",
    "B021": "挺立舒",     # 雙十節KF94 只是節慶單品名，不是品牌
    "B023": "萊潔",       # 傳說對決是遊戲聯名款式名，不是品牌
    "B029": "華淨",
    "B033": "聚泰",
    "B035": "優美特",
    "B036": "普潔",
    "B037": "佑合",
    "B041": "艾可兒",     # 跑跑薑餅人只是單品款式名
    "B043": "中衛",
    "B044": "幸福物語",
    "B050": "迪士尼",
    "B001": "BNN",
}


def brand_of(name):
    for b in BRANDS:
        if b in name:
            return b
    m = re.match(r"^[\[【〔]([^\]】〕]{1,6})", name)
    if m:
        return m.group(1)
    return re.split(r"[｜|／/ 　\-∣]", name)[0][:6]


def main():
    wb = openpyxl.load_workbook(SHOPEE, read_only=True, data_only=True)
    by3 = collections.defaultdict(collections.Counter)
    n_rows = 0
    for sheet in wb.sheetnames:
        if not sheet.startswith("📦 ") or sheet == "📦 商品列表":
            continue
        ws = wb[sheet]
        for r in ws.iter_rows(min_row=4, values_only=True):
            if not (r[1] and r[2]):
                continue
            sku, name = str(r[1]).strip(), str(r[2]).strip()
            m = re.match(r"^B(\d{3})\d", sku)
            if not m:
                continue
            by3["B" + m.group(1)][brand_of(name)] += 1
            n_rows += 1
    wb.close()

    codes = {}
    review = []
    for code, c in sorted(by3.items()):
        if code in OVERRIDES:
            codes[code] = OVERRIDES[code]
            continue
        top, n = c.most_common(1)[0]
        total = sum(c.values())
        codes[code] = top
        if n / total < 0.8 or total < 2:
            review.append({"code": code, "picked": top, "counts": dict(c.most_common(4))})

    out = {
        "_說明": "口罩貨號 3 碼公司碼(B+3位數字) → 品牌。可手改。keywords 是比對品名的兜底清單。",
        "codes": codes,
        "keywords": BRANDS,
        "_待人工複查": review,
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"掃了 {n_rows} 筆，寫出 {OUT}  ({len(codes)} 碼, {len(review)} 待複查)")


if __name__ == "__main__":
    main()
