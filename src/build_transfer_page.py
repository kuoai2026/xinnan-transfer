#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
產生「倉庫調貨單」門市端頁面 index.html。

資料來源：
  1. ~/我的雲端硬碟/產品進貨單/庫存盤點表/庫存盤點表_YYYY-MM-DD.xlsx （最新一份原始下載）
       欄位：主貨號 / 系統商品名 / 選項貨號 / 儲位 / 重量 / 規格一 / 規格二 / 可售庫存 / 總庫存 ...
  2. ~/我的雲端硬碟/蝦皮獲利計算表/蝦皮獲利計算表_含稅版_updated.xlsx → 「📦 商品列表」
       貨號 → 商品名稱（當別名／變體名兜底）
  3. transfer_page/brand_map.json —— 口罩貨號前4碼 → 品牌（make_brand_map.py 產、可手改）

輸出結構：
  口罩(B開頭，非BD) → 三層：品牌 > 子系列 > 變體
  其他分類          → 兩層：系列(清過的系統商品名) > 變體

設定檔：transfer_page/config.json
"""
import json, re, sys, os, glob, datetime

HERE = os.path.dirname(os.path.abspath(__file__))
HOME = os.path.expanduser("~")
INV_DIR = os.path.join(HOME, "我的雲端硬碟/產品進貨單/庫存盤點表")
SHOPEE_XLSX = os.path.join(HOME, "我的雲端硬碟/蝦皮獲利計算表/蝦皮獲利計算表_含稅版_updated.xlsx")
TEMPLATE = os.path.join(HERE, "template.html")
CONFIG = os.path.join(HERE, "config.json")
BRAND_MAP = os.path.join(HERE, "brand_map.json")

CAT_BY_PREFIX = [
    ("PW", "充電線材"), ("PG", "益智玩具"), ("BD", "入浴球"), ("NO", "入浴球"),
    ("A", "保健食品"), ("B", "醫療口罩"), ("C", "傷口護理"), ("D", "營養補充"),
    ("E", "嬰幼兒清潔"), ("F", "生活用品"), ("G", "3C配件"), ("H", "個人護理"),
    ("I", "食品飲品"), ("J", "成人用品"),
]

SEP = "∣｜|/／"
_SEPRE = re.compile("[" + re.escape(SEP) + "]")
# 變體/子系列名要清掉的雜訊
_NOISE = re.compile(
    r"(\(?\s*\d+\s*入(?:/\s*\d+\s*盒|/\s*盒|/\s*袋)?\s*\)?)"      # 50入 (50入/盒)
    r"|(（[^）]*贈[^）]*）)|(\[[^\]]*贈[^\]]*\])"                   # 贈品
    r"|(★[^★]*★?)|(☆[^☆]*☆?)|(⭐[^\s]*)"                          # 星星促銷
    r"|(【[^】]*】)"                                               # 剩餘方括號
    r"|(共\s*\d+\s*入)|(\d+\s*片裝)|(每包\s*\d+\s*片)"
    r"|(現貨(?:出清)?|台灣製(?:造)?|醫療口罩|醫用口罩|盒裝|袋裝|獨立包裝?)"
)
# 對象（依序比對，長/專一的在前）
TARGET_KW = ["嬰幼", "嫩嬰", "幼幼", "幼童", "小童", "中童", "大童", "兒童",
             "小顏", "小臉", "XL加大", "成人加大", "加大", "XL", "成人"]
TARGET_NORM = {"嫩嬰": "嬰幼", "小童": "幼童", "XL加大": "加大", "成人加大": "加大", "XL": "加大"}
# 款式（依序比對，複合詞在前）
STYLE_KW = ["一級薄", "一級厚", "二級薄", "二級厚", "一級", "二級",
            "全彩", "滿版", "活性碳", "不脫妝", "泡泡紋", "呼吸", "蝶形", "蝶型",
            "魚口", "KF94", "KN95", "N95", "鈔票", "4D", "3D", "立體", "平面"]
STYLE_NORM = {"蝶型": "蝶形", "3D": "立體", "4D": "立體"}
# 貨號第4碼（款式碼）兜底：規格一/品名判不到款式時才用這個補
# 例：B0041→1→平面、B0042→2→立體（郡昱呼吸口罩系列規格一完全沒寫立體/平面字樣，只能靠這個分辨）
DIGIT_STYLE = {"1": "平面", "2": "立體"}
# 這些「款式」本身不代表平面/立體（同名下兩種版型都有可能），判到時還要用貨號碼補上平面/立體
STYLE_AMBIGUOUS = {"呼吸", "鈔票", "全彩", "活性碳", "不脫妝"}
# 角色 / 聯名系列（規格一開頭 token 命中才當角色）
ROLE_KW = ["庫洛米", "酷洛米", "大耳狗", "美樂蒂", "布丁狗", "玉桂狗", "雙子星", "帕恰狗",
           "人魚漢頓", "korilakkuma", "拉拉熊", "柴語錄", "角落生物", "大甲媽",
           "寶可夢", "kitty", "KITTY", "凱蒂貓", "hello kitty", "史努比",
           "蠟筆小新", "卡娜赫拉", "吉伊卡哇", "小八貓", "烏薩奇"]

# 變體名要清掉的字（款式/尺寸/包裝/雜訊）
_VJUNK = re.compile(
    r"3D彈力|3D耳掛|3D耳繩|3D|4D"
    r"|適用?[0-9０-９\-~至\s]*歲(?:以[上下])?|[0-9０-９]+\s*[-~至]\s*[0-9０-９]+\s*歲|[0-9０-９]+\s*歲"
    r"|[0-9０-９]+\s*[.\-~至x*×＊]\s*[0-9０-９]+(?:\s*[.\-~至x*×＊]\s*[0-9０-９]+)?\s*(?:公分|cm)?"
    r"|共?\s*[0-9０-９]+\s*入|[0-9０-９]+\s*片(?:裝)?"
    r"|(?<![一-鿿])(XXS|XXL|XS|XL|[SMLＭＬＳ])(?![一-鿿a-z])"
    r"|立體|平面口罩|平面|滿版|全彩|活性碳|不脫妝|泡泡紋|呼吸口罩|呼吸"
    r"|蝶形|蝶型|魚口|KF94|KN95|N95|一級|二級|(?<![一-鿿])厚(?![一-鿿])|(?<![一-鿿])薄(?![一-鿿])"
    r"|口罩|醫療|醫用|盒裝|袋裝|(?<![一-鿿])盒(?![一-鿿])|(?<![一-鿿])袋(?![一-鿿])|款式|(?<![一-鿿])款(?![一-鿿])"
    r"|系列|版型|(?<![一-鿿])版(?![一-鿿])|(?<![一-鿿])型(?![一-鿿])|正反帶|細耳繩|耳繩|耳掛"
    r"|成人|兒童|幼幼|幼童|小童|中童|大童|嬰幼|嫩嬰|小顏|小臉|加大|婦幼|適小臉|小童適用|適用"
    r"|new|NEW|現貨(?:出清)?|限量|贈品|備註款式|直播限定|隨機出貨|不挑色|請選色|(?<![一-鿿])倒(?![一-鿿])"
    r"|(?:每[批包])?深淺略?(?:不同|隨機)|略不同|每[批包]深淺|略有深淺色?差|將停產|絕版|停產|新款上市|上市"
    r"|我愛媽祖|\{\{|\}\}|＿|一代|二代|獨立包裝?|禮盒?組?|(?<![一-鿿A-Za-z])DG"
    r"|[0-9０-９]+", re.I)

_VLEAD = re.compile(r"^[\s＿_\-－–—、,.。/／｜|∣:：;；()（）\[\]﹝﹞0-9０-９]+")
_VBRACKET = re.compile(r"[﹝﹞【】\[\]（）()]")


def strip_vjunk(s):
    s = _VBRACKET.sub(" ", s)
    # 兩次：像「5款共50入」這種，第一輪把「5」「共50入」清掉後「款」才會跟前後字斷開
    # （沒斷開時「款」的鄰接字元檢查會被相鄰雜訊卡住，鎖住不敢清），第二輪才清得掉。
    s = _VJUNK.sub(" ", s)
    s = _VJUNK.sub(" ", s)
    s = re.sub(r"[★☆⭐:：,\.、。/／\-－—_~♡❤｜|∣＊*]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = _VLEAD.sub("", s).strip()
    s = re.sub(r"^款\s+", "", s).strip()
    return s

_BRACKET = re.compile(r"[【\[（(〔「〈][^】\])）〕」〉]*[】\])）〕」〉]")


def log(*a):
    print(*a, file=sys.stderr)


def latest_inventory():
    pat = re.compile(r"庫存盤點表_(\d{4}-\d{2}-\d{2})\.xlsx$")
    best = None
    for p in glob.glob(os.path.join(INV_DIR, "庫存盤點表_*.xlsx")):
        m = pat.search(os.path.basename(p))
        if not m:
            continue
        if best is None or m.group(1) > best[1]:
            best = (p, m.group(1))
    if not best:
        sys.exit("找不到 庫存盤點表_YYYY-MM-DD.xlsx")
    return best


def category(sku):
    s = (sku or "").upper()
    for pre, name in CAT_BY_PREFIX:
        if s.startswith(pre):
            return name
    return "其他"


def norm(v):
    return "" if v is None else str(v).strip()


def clean(s):
    """去雜訊、壓空白、去頭尾符號"""
    s = _NOISE.sub(" ", s)
    s = re.sub(r"[()（）\[\]★☆⭐:：,\.、。/／\-—_~]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip(" ｜|∣/／-")
    return s.strip()


def lcp(strs):
    strs = [s for s in strs if s]
    if not strs:
        return ""
    a, b = min(strs), max(strs)
    i = 0
    while i < len(a) and i < len(b) and a[i] == b[i]:
        i += 1
    return a[:i]


def load_aliases():
    import openpyxl
    if not os.path.exists(SHOPEE_XLSX):
        log("警告：找不到蝦皮獲利計算表，略過別名")
        return {}
    wb = openpyxl.load_workbook(SHOPEE_XLSX, read_only=True, data_only=True)
    alias = {}

    def scan(ws):
        n = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            if len(row) < 3 or not row[1] or not row[2]:
                continue
            sku, name = str(row[1]).strip(), str(row[2]).strip()
            if sku and name and sku not in alias:
                alias[sku] = name
                n += 1
        return n

    main = wb["📦 商品列表"] if "📦 商品列表" in wb.sheetnames else wb[wb.sheetnames[0]]
    n_main = scan(main)
    if n_main == 0:
        # 「商品列表」的商品名稱欄有時是公式格，data_only 讀不到快取值（整欄變空）；
        # 各分類分頁（📦 A保健食品/📦 B醫療口罩…）是唯讀連動但存的是字面值，退回去掃那些。
        log("警告：📦 商品列表 商品名稱欄讀到空值（可能是公式沒快取），改掃各分類分頁")
        for name in wb.sheetnames:
            if name.startswith("📦 ") and name != "📦 商品列表":
                scan(wb[name])
    wb.close()
    return alias


def load_brand_map():
    if not os.path.exists(BRAND_MAP):
        log("警告：找不到 brand_map.json")
        return {"codes": {}, "aliases": {}, "keywords": []}
    with open(BRAND_MAP, encoding="utf-8") as f:
        d = json.load(f)
    return {"codes": d.get("codes", {}), "aliases": d.get("aliases", {}),
            "keywords": d.get("keywords", [])}


# ---------- 口罩：品牌 ----------
def mask_brand(sku, texts, bmap):
    code = sku[:4]  # "B"+3碼公司碼；第4碼是款式碼(1平面/2立體)，不分品牌
    b = bmap["codes"].get(code)
    if b and _clean_brand_ok(b):
        return b
    for kw in bmap["keywords"]:
        if any(kw in t for t in texts):
            return kw
    m = re.match(r"^([一-鿿A-Za-z]{2,6})", clean(texts[0]) if texts else "")
    return m.group(1) if m else "其他口罩"


def _clean_brand_ok(b):
    # 長度上限拉到 10：之前設 6 把「MissMix」(7字) 這種合法但較長的品牌名也擋掉了，
    # 導致 brand_map 裡手動 OVERRIDE 的值被自己的檢查否決，整組貨號退回去用花色名
    # 誤判成品牌（B0122 系列每個花色各自變成一張獨立品牌卡片）。
    return bool(b) and not re.search(r"[《》\d]", b) and len(b) <= 10 and b not in ("S", "M", "L")


def _pick_kw(texts, kws, norm_map):
    """在 texts 裡找 kws（依序），回傳最常出現的一個（正規化後）"""
    ct = {}
    for t in texts:
        for k in kws:
            if k.lower() in t.lower():
                nk = norm_map.get(k, k)
                ct[nk] = ct.get(nk, 0) + 1
                break
    if not ct:
        return ""
    return max(ct, key=lambda k: ct[k])


_KT_RE = re.compile(r"(?<![a-z])kt(?![a-z])", re.I)  # KT 是 Kitty 縮寫，要卡邊界避免亂中


def role_of(text):
    # 全字串找，不要只看第一段——蝦皮品名常用「-」當分隔符（水舞-成人/酷洛米一代-紫色），
    # 而 "-" 不在 _SEPRE 的分隔符清單裡，只看「第一個分隔符前」會漏掉角色字樣，
    # 之後清雜訊時就清不掉，殘留在變體名裡。
    t = (text or "").lower()
    for r in ROLE_KW:
        if r.lower() in t:
            return "KT" if r.lower() in ("kitty", "hello kitty", "凱蒂貓") else r
    if _KT_RE.search(t):
        return "KT"
    return ""


_KT_SPELLINGS = ["hello kitty", "kitty", "凱蒂貓"]


def variant_name(g1, strip_names, role, alias_name, sku):
    # 2026-09-05 使用者定案：直接用蝦皮獲利計算表品名(alias_name)，網翼規格一只當
    # alias 沒資料時的備援；不要再跑清雜訊/去角色/截斷那一整套——只去掉品牌字樣
    # （strip_names = 品牌 + 同公司碼底下出現過的其他品牌字樣，例：安心罩護是昌明的別名），
    # 因為品牌已經是卡片標題，重複顯示會很亂；其他字（對象/款式/角色/入數…）全部保留、不截斷。
    v = alias_name or g1 or ""
    for d in strip_names:
        if d:
            v = re.sub(re.escape(d), " ", v, flags=re.I)
    v = re.sub(r"[｜|∣]+", " ", v)
    v = re.sub(r"\s+", " ", v).strip(" ／/\\-－—_")
    return v or sku


def digit_style(sku):
    """貨號第4碼款式碼兜底：B0041→'1'→平面、B0042→'2'→立體"""
    m = re.match(r"^B\d{3}(\d)", sku)
    return DIGIT_STYLE.get(m.group(1), "") if m else ""


_COUNT_RE = re.compile(r"([0-9０-９]+)\s*入")


# ---------- 口罩：組三層（品牌 > 對象+款式+角色 > 顏色）----------
def build_mask_tree(rows, alias, bmap):
    brands = {}   # brand -> { linekey -> {parts, items} }
    brand_aliases = {}   # brand -> set(同公司碼底下出現過的其他品牌字樣，供搜尋用)
    for sku, sysname, g1, g2, avail in rows:
        al = alias.get(sku, "")
        brand = mask_brand(sku, [g1, sysname, al], bmap)
        # 對象/款式 只從規格一＋蝦皮品名（不碰 sysname SEO）；規格一判不到款式才靠貨號補
        target = _pick_kw([g1, al], TARGET_KW, TARGET_NORM) or "成人"
        style = _pick_kw([g1, al], STYLE_KW, STYLE_NORM)
        if style in ("一級", "二級"):
            # 厚/薄常寫在括號裡跟「一級/二級」不相鄰（如「一級醫療口罩(厚)」），
            # STYLE_KW 的「一級薄/一級厚」複合詞比對不到，要另外抓再拼起來
            for t in (g1, al):
                m = re.search(r"[(（](厚|薄)[)）]", t or "")
                if m:
                    style = style + m.group(1)
                    break
        ds = digit_style(sku)
        if not style:
            style = ds
        elif ds and style in STYLE_AMBIGUOUS and ds not in style:
            # 這些款式本身不代表平面/立體（呼吸/鈔票/全彩/活性碳/不脫妝都有兩種版型），
            # 規格一沒寫清楚時用貨號碼補上去，例：「呼吸」→「平面呼吸」
            style = ds + style
        role = role_of(g1) or role_of(al)
        lk = (brand, target, style, role)
        strip_names = bmap["aliases"].get(sku[:4], [brand])
        brand_aliases.setdefault(brand, set()).update(strip_names)
        d = brands.setdefault(brand, {}).setdefault(lk, {"items": []})
        d["items"].append((sku, sysname, g1, g2, avail, al, strip_names))

    out = []
    for brand, lines in brands.items():
        line_objs = []
        for (b, target, style, role), d in lines.items():
            items = d["items"]
            parts = [brand]
            if role:
                parts.append(role)
            if target and target != "成人":
                parts.append(target)
            elif target == "成人" and len(lines) > 1:
                parts.append("成人")
            if style:
                parts.append(style)
            name = " ".join(dict.fromkeys(parts))
            # 先算出每個變體的基本名 + 入數（若有），碰撞時優先用入數區分，
            # 而不是直接貼看不懂的貨號尾碼（同色不同入數是真的差異，不該被清掉）
            raw = []
            for sku, sysname, g1, g2, avail, al, strip_names in items:
                v = variant_name(g1, strip_names, role, al, sku)
                cm = _COUNT_RE.search(al or g1 or "")
                raw.append((sku, v, (cm.group(1) + "入") if cm else "", avail))
            base_ct = {}
            for _, v, _, _ in raw:
                base_ct[v] = base_ct.get(v, 0) + 1
            variants = []
            seen = set()
            for sku, v, cnt, avail in raw:
                nm = f"{v} {cnt}" if base_ct[v] > 1 and cnt else v
                if nm in seen:      # 入數也一樣才真的退回貨號尾碼
                    nm = f"{nm}·{sku[-3:]}"
                seen.add(nm)
                variants.append({"s": sku, "g": nm, "v": avail})
            line_objs.append({"n": name, "i": variants})
        line_objs.sort(key=lambda l: (-sum(1 for x in l["i"] if x["v"] > 0), l["n"]))
        cat = category(next(iter(lines.values()))["items"][0][0])
        # x：這個品牌底下曾出現過的其他品牌字樣（如「安心罩護」是「昌明」的別名）。
        # 顯示文字已經把這些字樣清掉了，但還是要留著給搜尋用，不然打「安心罩護」會搜不到。
        alias_words = sorted(w for w in brand_aliases.get(brand, ()) if w != brand)
        entry = {"t": "brand", "n": brand, "c": cat, "lines": line_objs}
        if alias_words:
            entry["x"] = " ".join(alias_words)
        out.append(entry)
    out.sort(key=lambda b: -sum(len(l["i"]) for l in b["lines"]))
    return out


# ---------- 其他分類：兩層（清過的系列名） ----------
_SEO_STOP = ["醫療口罩", "醫用口罩", "醫療級", "3D立體口罩", "台灣製造", "台灣製", "獨家",
             "限量", "現貨出清", "現貨", "免運", "優惠", "賣場", "搶先上市", "最新", "新款",
             "熱銷", "特價", "獨立包裝", "獨立包", "系列", "多款", "多色", "彩色", "節慶"]


def series_label(sysname):
    brands = re.findall(r"[【\[〔「]([^】\]〕」]{1,8})[】\]〕」]", sysname)
    brand = next((b for b in brands if not re.search(r"[《》!！\d]", b)
                  and "口罩" not in b and "聯名" not in b), "")
    s = _BRACKET.sub(" ", sysname)
    for w in _SEO_STOP:
        s = s.replace(w, " ")
    s = re.sub(r"[★☆｜|∣!！『』{}]", " ", s)
    toks = [t for t in re.split(r"[\s／/,、\-]+", s) if t and t not in ("入", "盒", "版", "型")]
    toks = list(dict.fromkeys(toks))
    label = " ".join(([brand] if brand else []) + toks[:4]).strip()
    label = re.sub(r"^[^\w一-鿿]+", "", re.sub(r"\s+", " ", label)).strip()
    if not label or len(label) > 24:
        cjk = re.findall(r"[一-鿿A-Za-z0-9]{2,10}", _BRACKET.sub(" ", sysname))
        label = " ".join(dict.fromkeys(([brand] if brand else []) + cjk[:2]))[:24] or sysname
    return label


def build_other(rows, alias):
    groups, order, sa = {}, [], {}
    for sku, sysname, g1, g2, avail in rows:
        disp = g1 or g2 or sysname
        if g1 and g2 and g2 not in ("-", ""):
            disp = f"{g1}／{g2}"
        if sysname not in groups:
            groups[sysname] = []
            sa[sysname] = set()
            order.append(sysname)
        groups[sysname].append({"s": sku, "g": disp, "v": avail})
        if alias.get(sku):
            sa[sysname].add(alias[sku])
    out = []
    for sysname in order:
        items = groups[sysname]
        # n 保留原始系統商品名給搜尋用（別再清乾淨——之前清乾淨過的 d 曾把「悠斯晶」這種
        # 品牌詞整個切掉，搜品牌名反而搜不到自己的商品）；d 才是清過的顯示用短名。
        entry = {"t": "series", "n": sysname, "d": series_label(sysname),
                 "c": category(items[0]["s"]), "i": items}
        ali = " ".join(sorted(sa[sysname]))[:300]
        if ali:
            entry["x"] = ali
        out.append(entry)
    return out


def load_and_build():
    import openpyxl
    inv_path, inv_date = latest_inventory()
    log(f"庫存來源：{os.path.basename(inv_path)}")
    alias = load_aliases()
    bmap = load_brand_map()
    log(f"別名 {len(alias)} 筆、品牌碼 {len(bmap['codes'])} 條")

    wb = openpyxl.load_workbook(inv_path, read_only=True, data_only=True)
    ws = wb["庫存盤點"] if "庫存盤點" in wb.sheetnames else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    idx = {n: i for i, n in enumerate(header)}

    def col(r, k):
        i = idx.get(k)
        return r[i] if i is not None and i < len(r) else None

    mask_rows, other_rows = [], []
    n = 0
    for r in it:
        sku = norm(col(r, "選項貨號"))
        sysname = norm(col(r, "系統商品名"))
        g1 = norm(col(r, "規格一"))
        g2 = norm(col(r, "規格二"))
        av = col(r, "可售庫存")
        if not sku or not sysname:
            continue
        try:
            av = int(av) if av is not None and str(av).strip() != "" else 0
        except (TypeError, ValueError):
            av = 0
        rec = (sku, sysname, g1, g2, av)
        if re.match(r"^B\d{4}", sku):
            mask_rows.append(rec)
        else:
            other_rows.append(rec)
        n += 1
    wb.close()

    tree = build_mask_tree(mask_rows, alias, bmap) + build_other(other_rows, alias)
    log(f"catalog：口罩 {len(mask_rows)} 品項 / 其他 {len(other_rows)} 品項 / "
        f"{sum(1 for x in tree if x['t']=='brand')} 品牌 + "
        f"{sum(1 for x in tree if x['t']=='series')} 系列")
    return tree, inv_date, n


def main():
    tree, inv_date, n_items = load_and_build()

    with open(CONFIG, encoding="utf-8") as f:
        cfg = json.load(f)
    config_json = {
        "endpoint": cfg.get("endpoint", ""),
        "token": cfg.get("token", ""),
        "persons": cfg.get("persons", ["瓊如", "柔柔", "阿霞", "怡虹", "品萱"]),
    }
    now = datetime.datetime.now()
    build_info = {"stock": inv_date, "built": f"{now:%m/%d %H:%M}"}

    with open(TEMPLATE, encoding="utf-8") as f:
        tpl = f.read()
    out = (tpl
           .replace("__CONFIG_JSON__", json.dumps(config_json, ensure_ascii=False))
           .replace("__BUILD_INFO_JSON__", json.dumps(build_info, ensure_ascii=False))
           .replace("__CATALOG_JSON__", json.dumps(tree, ensure_ascii=False, separators=(",", ":"))))

    repo_dir = os.path.expanduser(cfg.get("repo_dir", os.path.join(HOME, "repos/xinnan-transfer")))
    os.makedirs(repo_dir, exist_ok=True)
    out_path = os.path.join(repo_dir, "index.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(out)
    log(f"寫出 {out_path}  ({os.path.getsize(out_path)/1024:.0f} KB)")
    print(out_path)


if __name__ == "__main__":
    main()
